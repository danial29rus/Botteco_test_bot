from django.shortcuts import render
from django.http import HttpResponse
import openpyxl

from .models import Customer, Order

def export_to_excel(request):
    return render(request, 'export_to_excel.html')

def export_data(request):
    customers = Customer.objects.all()
    orders = Order.objects.all()

    wb = openpyxl.Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value="Customer Name")
    ws.cell(row=1, column=2, value="Address")
    ws.cell(row=1, column=3, value="Order Amount")

    row = 2
    for customer in customers:
        ws.cell(row=row, column=1, value=customer.name)
        ws.cell(row=row, column=2, value=customer.address)
        row += 1

    row = 2
    for order in orders:
        ws.cell(row=row, column=3, value=order.amount)
        row += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=data.xlsx'
    wb.save(response)

    return response
